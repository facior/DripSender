"""Wczytywanie listy odbiorców z pliku CSV lub Excela."""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

MAX_PODGLAD = 8
MAX_WIERSZY = 50000

try:  # pragma: no cover - opcjonalna zależność
    from openpyxl import load_workbook

    _HAS_EXCEL = True
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]
    _HAS_EXCEL = False

# Nagłówki, po których zgadujemy przeznaczenie kolumny.
PODPOWIEDZI = {
    "email": ("email", "e-mail", "mail", "adres", "adres e-mail", "poczta"),
    "imie": ("imie", "imię", "name", "first name", "osoba", "kontakt"),
    "firma": ("firma", "company", "nazwa", "nazwa firmy", "klient", "organizacja"),
}


def excel_available() -> bool:
    return _HAS_EXCEL


@dataclass
class Arkusz:
    """Surowa zawartość pliku: nagłówki i wiersze."""

    naglowki: list[str] = field(default_factory=list)
    wiersze: list[list[str]] = field(default_factory=list)
    zrodlo: str = ""

    @property
    def podglad(self) -> list[list[str]]:
        return self.wiersze[:MAX_PODGLAD]

    def zgadnij_mapowanie(self, pola: list[str]) -> dict[str, int]:
        """Dopasowuje kolumny pliku do pól programu na podstawie nagłówków."""
        mapowanie: dict[str, int] = {}
        znormalizowane = [(h or "").strip().lower() for h in self.naglowki]

        for docelowe in ["email", *pola]:
            podpowiedzi = PODPOWIEDZI.get(docelowe, (docelowe,))
            for index, naglowek in enumerate(znormalizowane):
                if index in mapowanie.values():
                    continue
                if naglowek in podpowiedzi or any(p in naglowek for p in podpowiedzi):
                    mapowanie[docelowe] = index
                    break

        if "email" not in mapowanie:
            # Bez pasującego nagłówka szukamy kolumny, która wygląda na adresy.
            for index in range(len(self.naglowki)):
                wartosci = [w[index] for w in self.wiersze[:20] if index < len(w)]
                if wartosci and sum("@" in v for v in wartosci) >= max(1, len(wartosci) // 2):
                    mapowanie["email"] = index
                    break
        return mapowanie


def wczytaj(sciezka: str | Path, ma_naglowki: bool = True) -> Arkusz:
    """Wczytuje CSV albo XLSX. Podnosi ValueError z czytelnym komunikatem."""
    sciezka = Path(sciezka)
    if not sciezka.is_file():
        raise ValueError("Nie znaleziono pliku: " + str(sciezka))

    if sciezka.suffix.lower() in (".xlsx", ".xlsm"):
        return _wczytaj_excel(sciezka, ma_naglowki)
    return _wczytaj_csv(sciezka, ma_naglowki)


def _wczytaj_csv(sciezka: Path, ma_naglowki: bool) -> Arkusz:
    surowe = None
    for kodowanie in ("utf-8-sig", "cp1250", "latin-2"):
        try:
            surowe = sciezka.read_text(encoding=kodowanie)
            break
        except UnicodeDecodeError:
            continue
    if surowe is None:
        raise ValueError("Nie udało się odczytać pliku - nieznane kodowanie znaków.")

    probka = surowe[:4000]
    try:
        dialekt = csv.Sniffer().sniff(probka, delimiters=";,\t|")
        separator = dialekt.delimiter
    except csv.Error:
        separator = ";" if probka.count(";") >= probka.count(",") else ","

    wiersze = [
        [(k or "").strip() for k in wiersz]
        for wiersz in csv.reader(surowe.splitlines(), delimiter=separator)
        if any((k or "").strip() for k in wiersz)
    ]
    return _zbuduj(wiersze, ma_naglowki, sciezka.name + "  (separator: " + separator + ")")


def _wczytaj_excel(sciezka: Path, ma_naglowki: bool) -> Arkusz:
    if not _HAS_EXCEL:
        raise ValueError(
            "Obsługa Excela wymaga biblioteki openpyxl. Zapisz plik jako CSV "
            "albo doinstaluj bibliotekę."
        )
    try:
        skoroszyt = load_workbook(sciezka, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Nie udało się otworzyć arkusza: " + str(exc)) from exc

    arkusz = skoroszyt.active
    wiersze: list[list[str]] = []
    for wiersz in arkusz.iter_rows(values_only=True):
        komorki = ["" if k is None else str(k).strip() for k in wiersz]
        if any(komorki):
            wiersze.append(komorki)
        if len(wiersze) >= MAX_WIERSZY:
            break
    skoroszyt.close()
    return _zbuduj(wiersze, ma_naglowki, sciezka.name + "  (arkusz: " + str(arkusz.title) + ")")


def _zbuduj(wiersze: list[list[str]], ma_naglowki: bool, zrodlo: str) -> Arkusz:
    if not wiersze:
        raise ValueError("Plik nie zawiera żadnych danych.")
    szerokosc = max(len(w) for w in wiersze)
    wyrownane = [w + [""] * (szerokosc - len(w)) for w in wiersze]

    if ma_naglowki:
        naglowki = [h or ("Kolumna " + str(i + 1)) for i, h in enumerate(wyrownane[0])]
        dane = wyrownane[1:]
    else:
        naglowki = ["Kolumna " + str(i + 1) for i in range(szerokosc)]
        dane = wyrownane
    if not dane:
        raise ValueError("Plik zawiera tylko nagłówki, bez wierszy z danymi.")
    return Arkusz(naglowki=naglowki, wiersze=dane[:MAX_WIERSZY], zrodlo=zrodlo)


def na_odbiorcow(
    arkusz: Arkusz, mapowanie: dict[str, int], pola: list[str]
) -> list[tuple[str, dict[str, str]]]:
    """Przekłada wiersze na pary (email, dane) gotowe dla bazy."""
    kolumna_email = mapowanie.get("email")
    if kolumna_email is None:
        raise ValueError("Nie wskazano kolumny z adresem e-mail.")

    wynik: list[tuple[str, dict[str, str]]] = []
    for wiersz in arkusz.wiersze:
        adres = wiersz[kolumna_email] if kolumna_email < len(wiersz) else ""
        dane = {}
        for klucz in pola:
            index = mapowanie.get(klucz)
            if index is not None and index < len(wiersz):
                dane[klucz] = wiersz[index]
        wynik.append((adres, dane))
    return wynik
